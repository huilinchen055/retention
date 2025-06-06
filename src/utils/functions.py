"""FUNCTIONS MODULE"""

import random
import math
import json
import os
import requests
import urllib.parse
import pandas as pd
import numpy as np
from tqdm.notebook import tqdm
from shapely.geometry import Polygon, Point
from scipy.stats import gaussian_kde, truncnorm
import plotly.express as px
from imblearn.over_sampling import SMOTENC
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# custom modules
from src.utils.settings import *


def map_categories(identifier, key):
    """
    Assign the probabilities to the mapping
    """
    codes = CATEGORICAL_CUSTOM_MAPPINGS[identifier][key]["codes"]
    return np.random.choice(codes)


def fill_categorical_custom_mapping(identifier, prob_dist):
    """
    Assign the probabilities to the mapping
    """
    for key in prob_dist.keys():
        CATEGORICAL_CUSTOM_MAPPINGS[identifier][key]["p"] = prob_dist[key]


def aggregate_common_keys(data):
    """
    Aggregate probabiltites for keys sharing same codes in the mapping
    """
    lookup_dict = {}
    for item in data:
        codes = item["codes"]
        p = item["p"]
        for code in codes:
            if code in lookup_dict:
                lookup_dict[code] += p
            else:
                lookup_dict[code] = p

    result = [{"code": k, "p": v} for k, v in lookup_dict.items()]
    return sorted(result, key=lambda k: k["code"])


def generate_random_points_in_polygon(
    polygon, num_points, geoid, water_data, water_sindex
):
    """
    Generates random points intersecting within a polygon's boundary, excluding water bodies.
    Uses spatial indexing for efficiency.
    """
    points = []
    min_x, min_y, max_x, max_y = polygon.bounds

    count = 0
    while len(points) < num_points:
        random_point = Point(random.uniform(min_x, max_x), random.uniform(min_y, max_y))
        if random_point.within(polygon) and not any(
            random_point.within(water_data.iloc[i].geometry)
            for i in water_sindex.intersection(random_point.bounds)
        ):
            points.append((random_point, geoid))
        else:
            count += 1
        if count > 20000:
            print(f"Skipping zipcode: {geoid} | points: {len(points)} < {num_points}")
            break
    return points


def sample_coordinates(df, column, n=1, water_data=None):
    """
    Samples random coordinates within a geoid geometry based on its population
    and land area.
    """
    total_population = df[column].sum()
    max_distance_to_coast = max(df["MeanDistanceToCoast"])
    coordinates = []
    water_sindex = water_data.sindex
    for _, row in tqdm(df.iterrows(), total=df.shape[0]):
        population = row[column]
        distance_to_coast = row["MeanDistanceToCoast"]
        geoid = row["ZipCode"]

        # Adjust points by population density to consider land area
        points_by_population_density = max(
            5,
            math.ceil(
                (population / total_population) * n
            ),
        )
        points_by_dtc_gradient = max(
            5, 2 * math.ceil((max_distance_to_coast - distance_to_coast))
        )
        num_points = math.ceil(
            (points_by_population_density * 0.8) + (points_by_dtc_gradient * 0.2)
        )
        if isinstance(row["geometry"], Polygon):
            points = generate_random_points_in_polygon(
                row["geometry"], num_points, geoid, water_data, water_sindex
            )
            coordinates.extend([(point.y, point.x, geoid) for point, geoid in points])
    return coordinates


def sample_continuous_variable(df, var, smean=None, svariance=None):
    """
    Samples any continuous variable based on kernel density estimation and truncated normal distribution
    """
    data = df.loc[
        df[var] > 0, var
    ]  # business decision - we don't want 0 samples in proforma of continuous variable

    # fit a Gaussian KDE to the data
    kde = gaussian_kde(data)

    # determine the bounds of the truncated distribution
    lower_bound = min(data)
    upper_bound = max(data)
    mean, variance = kde.dataset.mean(), kde.dataset.var()
    if smean:
        mean += smean * mean
    if svariance:
        variance += svariance * variance

    a, b = (lower_bound - mean) / np.sqrt(variance), (upper_bound - mean) / np.sqrt(
        variance
    )

    # create a truncated normal distribution with the same mean and variance as the KDE
    truncnorm_dist = truncnorm(a, b, loc=mean, scale=np.sqrt(variance))

    # generate samples from the truncated distribution
    samples = truncnorm_dist.rvs(TOTAL_SAMPLES)

    samples = np.round(samples).astype(int)

    return samples


def adjust_probability_distribution(prob_list):
    """
    Distribute the probabilites for underrepresented key
    """
    max_prob = max(prob_list)
    max_prob_index = prob_list.index(max_prob)
    for index, prob in enumerate(prob_list):
        if prob < 0.01:
            prob_list[index] += MINIMUM_VAR_PROBABILITY_THRESHOLD
            prob_list[max_prob_index] -= MINIMUM_VAR_PROBABILITY_THRESHOLD
    return prob_list


def balance_probability_distribution(identifier, prob_map):
    """
    Divide the probabilities evenly for keys with multiple codes in it
    """
    for item in CATEGORICAL_CUSTOM_MAPPINGS[identifier].values():
        codes = item["codes"]
        for code in codes:
            lookup_dict = list(filter(lambda k: k["code"] == code, prob_map))[0]
            lookup_dict["p"] /= len(codes)
    return prob_map


def plot_pie(data, title):
    """Plot piechart for data"""
    dist = data.value_counts(normalize=True) * 100
    fig = px.pie(dist, values=dist.values, names=dist.keys(), title=title)
    fig.update_layout(title_x=0.5)
    return fig


def get_quote(df, policy_number, environment="stage"):
    px_central_api = (
        SAGESURE_ENVIRONMENT_URLS.get(environment)
        + "/cru-4/pxcentral/api/rest/v1/policies/"
        + policy_number
    )
    try:
        response = requests.get(
            px_central_api,
            auth=(SAGESURE_API_AUTH_USER, SAGESURE_API_AUTH_PASS),
            timeout=60,
        )
        response.raise_for_status()

        policy_xml = response.content

        # Write to a file instead of storing in memory
        file_path = f"../outputs/policies/{policy_number}.xml"
        os.makedirs(
            os.path.dirname(file_path), exist_ok=True
        )  # Ensure directory exists
        with open(file_path, "wb") as file:
            file.write(policy_xml)

        # Update DataFrame status
        df.loc[df["PolicyNumber"] == policy_number, "Fetched"] = True
    except requests.exceptions.RequestException as e:
        print(f"Failed to fetch policy {policy_number}: {e}")


def generate_smote_samples(X, Y, col_name):
    # Apply SMOTENC with categorical feature mask
    categorical_cols = X.select_dtypes(include=["object"]).columns.tolist()
    categorical_features_mask = [
        any(col.startswith(feature) for feature in categorical_cols)
        for col in X.columns
    ]

    # Apply SMOTENC
    smotenc = SMOTENC(
        k_neighbors=min(Y.value_counts()[-1] - 1, 5),
        random_state=42,
        categorical_features=categorical_features_mask,
    )

    # Use SMOTENC to generate synthetic samples based on X
    X_resampled, y_resampled = smotenc.fit_resample(X, Y)
    resampled_df = pd.concat(
        [X_resampled, pd.DataFrame(y_resampled, columns=[col_name])], axis=1
    )

    resampled_df.reset_index(inplace=True, drop=True)
    # Append only the synthetic samples (excluding the original base_data)
    synthetic_samples = resampled_df[len(X) :]

    return synthetic_samples


def generate_smote_samples_dummy_target(base_data, dummy_target):
    # Apply SMOTENC with categorical feature mask
    categorical_cols = base_data.select_dtypes(include=["object"]).columns.tolist()
    categorical_features_mask = [
        any(col.startswith(feature) for feature in categorical_cols)
        for col in base_data.columns
    ]

    # Apply SMOTENC
    smotenc = SMOTENC(
        k_neighbors=5, random_state=42, categorical_features=categorical_features_mask
    )

    # Use SMOTENC to generate synthetic samples based on base_data
    X_resampled, _ = smotenc.fit_resample(base_data, dummy_target)

    # Append only the synthetic samples (excluding the original base_data)
    synthetic_samples = X_resampled[len(base_data) :]

    return synthetic_samples


def subset_df(df1, df2, col_name, custom_prob):
    # Calculate the current distribution of df1
    current_dist = df1[col_name].value_counts(normalize=True).to_dict()

    # Determine the total number of samples needed
    total_samples = len(df1)

    # Calculate the number of samples needed from df2 for each category
    samples_needed = {}
    for category, desired_prob in custom_prob.items():
        current_count = current_dist.get(category, 0) * total_samples
        desired_count = desired_prob * total_samples
        samples_needed[category] = max(0, desired_count - current_count)

    # Sample the required number of rows from df2 for each category
    sampled_dfs = []
    for category, count in samples_needed.items():
        sample = df2[df2[col_name] == category].sample(n=int(count), replace=True)
        sampled_dfs.append(sample)

    # Concatenate the sampled dataframes
    final_sampled_df = pd.concat(sampled_dfs, axis=0).reset_index(drop=True)

    return final_sampled_df


def geocode(address):
    """Geocode an address to lat/lon"""
    try:
        final_url = f"{MAPRISK_GEOCODE_URL}{urllib.parse.quote_plus(address)}"
        raw_results = requests.get(final_url, timeout=10)
        results = raw_results.json()
        if results["success"] is True:
            lat = results["response"]["geocodeResults"][0]["latitude"]
            long = results["response"]["geocodeResults"][0]["longitude"]
            return lat, long
        else:
            print(results)
            raise Exception("Failed to get elevation")
            # return None, None
    except (TimeoutError, IndexError, TypeError):
        print(f"Failed to get geocode for {address}")
        return None, None


def get_elevation(lat, lon):
    """Get the elevation from lat/lon"""
    try:
        final_url = f"{MAPRISK_REPORTS_URL}{MAPRISK_REPORTSLIST[0]}&format=json&poi[latitude]={lat}&poi[longitude]={lon}"
        raw_results = requests.get(final_url, timeout=10)
        results = raw_results.json()
        if results["success"] is True:
            elevation = results["response"]["reportResults"]["elevation"]["elevation"]
            return elevation
        else:
            print(results)
            raise Exception("Failed to get elevation")
            # return None
    except (TimeoutError, IndexError, TypeError):
        print(f"Failed to get elevation for {lat}, {lon}")
        return None




def write_data(wb, sheet_name, df, start_row, start_col):
    """
    Writes the data from a pandas DataFrame to a specific location in an Excel sheet.

    Parameters:
    - wb: openpyxl Workbook object
    - sheet_name: string, name of the sheet to write data into
    - df: pandas DataFrame, the data to write
    - start_row: int, the starting row in the Excel sheet (1-indexed)
    - start_col: int or string, the starting column in the Excel sheet (1-indexed or column letter)
    """
    # Check if sheet exists, otherwise add it
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Convert start_col to a number if it's a letter
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)

    # Use dataframe_to_rows to convert DataFrame to rows suitable for openpyxl
    for r_index, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=start_row
    ):
        for c_index, value in enumerate(row, start=start_col):
            ws.cell(row=r_index, column=c_index, value=value)



def get_occupancy_class_code_group():
    """
    Get the occupancy class code group based on the occupancy class code
    """
    # Read the Excel file starting from row 7, columns B and F
    df = pd.read_excel("../data/local/Class code/SageSure BOP MASTER Class List v2.1.xlsx", sheet_name="Complete", skiprows=5, usecols="B,F")
    occupancy_mapping = {}

    # Create a mapping of occupancy type to class codes
    for _, row in df.iterrows():
        class_code = str(row["Class Code"])
        occupancy_type = str(row["Occupancy Class Type"])
        if occupancy_type not in occupancy_mapping:
            occupancy_mapping[occupancy_type] = []
        if class_code not in occupancy_mapping[occupancy_type]:  # Check if class_code is not already in the list
            occupancy_mapping[occupancy_type].append(class_code)

    return occupancy_mapping
